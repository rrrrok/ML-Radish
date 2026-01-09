"""
무 도매가 예측 프로젝트 - 고급 데이터 전처리 스크립트
전일, 전년 데이터 활용 및 휴일 처리 포함
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import warnings
try:
    import holidays
    HAS_HOLIDAYS = True
except ImportError:
    HAS_HOLIDAYS = False
    print("경고: holidays 라이브러리가 설치되지 않았습니다. 'pip install holidays'로 설치하세요.")
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("경고: openpyxl 라이브러리가 설치되지 않았습니다. 'pip install openpyxl'로 설치하세요.")
warnings.filterwarnings('ignore')


class AdvancedDataPreprocessor:
    """고급 데이터 전처리 클래스"""
    
    def __init__(self, price_file='price_data_original.xlsx', quantity_file='quantity_data_original.xlsx'):
        """
        초기화
        Args:
            price_file: 가격 데이터 파일 경로
            quantity_file: 반입량 데이터 파일 경로
        """
        self.price_file = price_file
        self.quantity_file = quantity_file
        self.df_price = None
        self.df_quantity = None
        
        # 한국 공휴일 객체 생성 (holidays 라이브러리 사용)
        if HAS_HOLIDAYS:
            # 데이터 범위를 고려하여 2018-2026년 공휴일 로드
            self.kr_holidays = holidays.Korea(years=range(2018, 2027))
        else:
            # 라이브러리가 없으면 빈 딕셔너리 사용
            self.kr_holidays = {}
    
    def _is_holiday(self, date):
        """날짜가 공휴일인지 확인 (holidays 라이브러리 사용)"""
        if not HAS_HOLIDAYS:
            return False
        
        # pd.Timestamp를 date 객체로 변환
        if isinstance(date, pd.Timestamp):
            date_obj = date.date()
        elif isinstance(date, datetime):
            date_obj = date.date()
        else:
            date_obj = pd.Timestamp(date).date()
        
        return date_obj in self.kr_holidays
    
    def _is_weekend(self, date):
        """날짜가 주말인지 확인"""
        if isinstance(date, pd.Timestamp):
            return date.weekday() >= 5
        return pd.Timestamp(date).weekday() >= 5
    
    def _is_trading_day(self, date):
        """거래일인지 확인 (주말, 공휴일 제외)"""
        return not (self._is_weekend(date) or self._is_holiday(date))
    
    def load_data(self):
        """데이터 로드 및 기본 전처리"""
        print("=" * 60)
        print("데이터 로드 중...")
        print("=" * 60)
        
        self.df_price = pd.read_excel(self.price_file)
        self.df_quantity = pd.read_excel(self.quantity_file)
        
        # DATE 컬럼을 datetime으로 변환
        self.df_price['DATE'] = pd.to_datetime(self.df_price['DATE'])
        self.df_quantity['DATE'] = pd.to_datetime(self.df_quantity['DATE'])
        
        # 날짜순 정렬
        self.df_price = self.df_price.sort_values('DATE').reset_index(drop=True)
        self.df_quantity = self.df_quantity.sort_values('DATE').reset_index(drop=True)
        
        # 비어있는 날짜 채우기
        self._fill_missing_dates()
        
        # 주말/공휴일 플래그 추가
        self.df_price['is_weekend'] = self.df_price['DATE'].apply(self._is_weekend)
        self.df_price['is_holiday'] = self.df_price['DATE'].apply(self._is_holiday)
        self.df_price['is_trading_day'] = self.df_price['DATE'].apply(self._is_trading_day)
        
        self.df_quantity['is_weekend'] = self.df_quantity['DATE'].apply(self._is_weekend)
        self.df_quantity['is_holiday'] = self.df_quantity['DATE'].apply(self._is_holiday)
        self.df_quantity['is_trading_day'] = self.df_quantity['DATE'].apply(self._is_trading_day)
        
        print(f"가격 데이터: {self.df_price.shape}")
        print(f"반입량 데이터: {self.df_quantity.shape}")
        print(f"가격 데이터 날짜 범위: {self.df_price['DATE'].min()} ~ {self.df_price['DATE'].max()}")
        print(f"반입량 데이터 날짜 범위: {self.df_quantity['DATE'].min()} ~ {self.df_quantity['DATE'].max()}")
    
    def _fill_missing_dates(self):
        """비어있는 날짜 채우기"""
        print("\n비어있는 날짜 채우기 중...")
        
        # 가격 데이터: 각 그룹별로 날짜 채우기
        group_cols = [col for col in self.df_price.columns 
                     if col not in ['DATE'] and self.df_price[col].dtype not in [np.int64, np.float64]]
        
        if len(group_cols) > 0:
            date_range = pd.date_range(
                start=self.df_price['DATE'].min(),
                end=self.df_price['DATE'].max(),
                freq='D'
            )
            
            # 각 그룹 조합에 대해 날짜 채우기
            all_combinations = self.df_price[group_cols].drop_duplicates()
            new_rows = []
            
            for _, combo in all_combinations.iterrows():
                group_data = self.df_price.copy()
                for g in group_cols:
                    group_data = group_data[group_data[g] == combo[g]]
                
                existing_dates = set(group_data['DATE'].dt.date)
                
                for date in date_range:
                    if date.date() not in existing_dates:
                        new_row = combo.to_dict()
                        new_row['DATE'] = date
                        # 숫자 컬럼은 NaN으로
                        for col in self.df_price.select_dtypes(include=[np.number]).columns:
                            new_row[col] = np.nan
                        new_rows.append(new_row)
            
            if new_rows:
                new_df = pd.DataFrame(new_rows)
                self.df_price = pd.concat([self.df_price, new_df], ignore_index=True)
                self.df_price = self.df_price.sort_values('DATE').reset_index(drop=True)
                print(f"  가격 데이터: {len(new_rows)}개 날짜 추가")
        
        # 반입량 데이터: 날짜 채우기
        group_cols_qty = [col for col in self.df_quantity.columns 
                         if col not in ['DATE'] and self.df_quantity[col].dtype not in [np.int64, np.float64]]
        
        if len(group_cols_qty) > 0:
            date_range = pd.date_range(
                start=self.df_quantity['DATE'].min(),
                end=self.df_quantity['DATE'].max(),
                freq='D'
            )
            
            all_combinations = self.df_quantity[group_cols_qty].drop_duplicates()
            new_rows = []
            
            for _, combo in all_combinations.iterrows():
                group_data = self.df_quantity.copy()
                for g in group_cols_qty:
                    group_data = group_data[group_data[g] == combo[g]]
                
                existing_dates = set(group_data['DATE'].dt.date)
                
                for date in date_range:
                    if date.date() not in existing_dates:
                        new_row = combo.to_dict()
                        new_row['DATE'] = date
                        for col in self.df_quantity.select_dtypes(include=[np.number]).columns:
                            new_row[col] = np.nan
                        new_rows.append(new_row)
            
            if new_rows:
                new_df = pd.DataFrame(new_rows)
                self.df_quantity = pd.concat([self.df_quantity, new_df], ignore_index=True)
                self.df_quantity = self.df_quantity.sort_values('DATE').reset_index(drop=True)
                print(f"  반입량 데이터: {len(new_rows)}개 날짜 추가")
        
    def identify_missing_values(self):
        """결측치 확인"""
        print("\n" + "=" * 60)
        print("결측치 분석")
        print("=" * 60)
        
        # 가격 데이터의 0 값 확인
        price_cols = self.df_price.select_dtypes(include=[np.number]).columns
        price_cols = [col for col in price_cols if col not in ['is_weekend', 'is_holiday', 'is_trading_day']]
        
        print("\n[가격 데이터] 0 값 개수:")
        for col in price_cols:
            zero_count = (self.df_price[col] == 0).sum()
            if zero_count > 0:
                print(f"  {col}: {zero_count}개 ({zero_count/len(self.df_price)*100:.2f}%)")
        
        # 반입량 데이터의 0 값 확인
        quantity_cols = self.df_quantity.select_dtypes(include=[np.number]).columns
        quantity_cols = [col for col in quantity_cols if col not in ['is_weekend', 'is_holiday', 'is_trading_day']]
        
        print("\n[반입량 데이터] 0 값 개수:")
        for col in quantity_cols:
            zero_count = (self.df_quantity[col] == 0).sum()
            if zero_count > 0:
                print(f"  {col}: {zero_count}개 ({zero_count/len(self.df_quantity)*100:.2f}%)")
    
    def fill_price_with_previous_day(self):
        """가격 데이터: 전일 데이터로 채우기"""
        print("\n" + "=" * 60)
        print("가격 데이터: 전일 데이터로 채우기")
        print("=" * 60)
        
        price_cols = [col for col in self.df_price.select_dtypes(include=[np.number]).columns 
                     if col not in ['is_weekend', 'is_holiday', 'is_trading_day']]
        group_cols = [col for col in self.df_price.columns 
                     if col not in price_cols and col not in ['DATE', 'is_weekend', 'is_holiday', 'is_trading_day']]
        
        # 0 값을 NaN으로 변환
        for col in price_cols:
            self.df_price.loc[self.df_price[col] == 0, col] = np.nan
        
        # 그룹별로 전일 데이터로 채우기
        for col in price_cols:
            initial_missing = self.df_price[col].isna().sum()
            
            if len(group_cols) > 0:
                # 같은 그룹 내에서 전일 값으로 채우기
                self.df_price[col] = self.df_price.groupby(group_cols)[col].transform(
                    lambda x: x.ffill()
                )
            
            filled = initial_missing - self.df_price[col].isna().sum()
            if filled > 0:
                print(f"  {col}: 전일 데이터로 {filled}개 채움")
    
    def fill_price_with_previous_year(self):
        """가격 데이터: 전년 동일일자 데이터로 채우기"""
        print("\n" + "=" * 60)
        print("가격 데이터: 전년 동일일자 데이터로 채우기")
        print("=" * 60)
        
        price_cols = [col for col in self.df_price.select_dtypes(include=[np.number]).columns 
                     if col not in ['is_weekend', 'is_holiday', 'is_trading_day']]
        group_cols = [col for col in self.df_price.columns 
                     if col not in price_cols and col not in ['DATE', 'is_weekend', 'is_holiday', 'is_trading_day']]
        
        for col in price_cols:
            initial_missing = self.df_price[col].isna().sum()
            if initial_missing == 0:
                continue
            
            # 전년 동일일자 데이터로 채우기
            filled_count = 0
            for idx, row in self.df_price[self.df_price[col].isna()].iterrows():
                current_date = row['DATE']
                prev_year_date = current_date - pd.DateOffset(years=1)
                
                # 같은 그룹의 전년 동일일자 데이터 찾기
                if len(group_cols) > 0:
                    group_values = {g: row[g] for g in group_cols}
                    prev_year_data = self.df_price[
                        (self.df_price['DATE'] == prev_year_date) &
                        (self.df_price[col].notna())
                    ]
                    
                    for g in group_cols:
                        prev_year_data = prev_year_data[prev_year_data[g] == group_values[g]]
                    
                    if len(prev_year_data) > 0 and prev_year_data[col].iloc[0] > 0:
                        self.df_price.loc[idx, col] = prev_year_data[col].iloc[0]
                        filled_count += 1
                else:
                    prev_year_data = self.df_price[
                        (self.df_price['DATE'] == prev_year_date) &
                        (self.df_price[col].notna())
                    ]
                    if len(prev_year_data) > 0 and prev_year_data[col].iloc[0] > 0:
                        self.df_price.loc[idx, col] = prev_year_data[col].iloc[0]
                        filled_count += 1
            
            if filled_count > 0:
                print(f"  {col}: 전년 동일일자 데이터로 {filled_count}개 채움")
    
    def fill_price_holiday_values(self):
        """가격 데이터: 휴일 처리 (선형보간)"""
        print("\n" + "=" * 60)
        print("가격 데이터: 휴일 처리 (선형보간)")
        print("=" * 60)
        
        price_cols = [col for col in self.df_price.select_dtypes(include=[np.number]).columns 
                     if col not in ['is_weekend', 'is_holiday', 'is_trading_day']]
        group_cols = [col for col in self.df_price.columns 
                     if col not in price_cols and col not in ['DATE', 'is_weekend', 'is_holiday', 'is_trading_day']]
        
        for col in price_cols:
            # 휴일인데 값이 없는 경우 선형보간으로 채우기
            holiday_mask = (~self.df_price['is_trading_day']) & (self.df_price[col].isna())
            initial_missing = holiday_mask.sum()
            
            if initial_missing == 0:
                continue
            
            # 그룹별로 선형보간 수행
            if len(group_cols) > 0:
                self.df_price[col] = self.df_price.groupby(group_cols)[col].transform(
                    lambda x: x.interpolate(method='linear', limit_direction='both')
                )
            else:
                self.df_price[col] = self.df_price[col].interpolate(method='linear', limit_direction='both')
            
            filled = initial_missing - self.df_price[col].isna().sum()
            if filled > 0:
                print(f"  {col}: 휴일 선형보간으로 {filled}개 채움")
    
    def fill_price_remaining_missing(self):
        """가격 데이터: 남은 결측치 처리 (보간 및 평균)"""
        print("\n" + "=" * 60)
        print("가격 데이터: 남은 결측치 처리")
        print("=" * 60)
        
        price_cols = [col for col in self.df_price.select_dtypes(include=[np.number]).columns 
                     if col not in ['is_weekend', 'is_holiday', 'is_trading_day']]
        group_cols = [col for col in self.df_price.columns 
                     if col not in price_cols and col not in ['DATE', 'is_weekend', 'is_holiday', 'is_trading_day']]
        
        for col in price_cols:
            initial_missing = self.df_price[col].isna().sum()
            if initial_missing == 0:
                continue
            
            # 방법 1: 선형 보간
            if len(group_cols) > 0:
                self.df_price[col] = self.df_price.groupby(group_cols)[col].transform(
                    lambda x: x.interpolate(method='linear', limit_direction='both')
                )
            
            # 방법 2: 이동평균
            if len(group_cols) > 0:
                self.df_price[col] = self.df_price.groupby(group_cols)[col].transform(
                    lambda x: x.fillna(x.rolling(window=5, center=True, min_periods=1).mean())
                )
            
            # 방법 3: 같은 요일의 평균
            self.df_price['weekday'] = self.df_price['DATE'].dt.dayofweek
            if len(group_cols) > 0:
                weekday_mean = self.df_price.groupby(['weekday'] + group_cols)[col].transform('mean')
                self.df_price[col] = self.df_price[col].fillna(weekday_mean)
            else:
                weekday_mean = self.df_price.groupby('weekday')[col].transform('mean')
                self.df_price[col] = self.df_price[col].fillna(weekday_mean)
            self.df_price = self.df_price.drop('weekday', axis=1)
            
            # 방법 4: 그룹별 평균
            if len(group_cols) > 0:
                group_mean = self.df_price.groupby(group_cols)[col].transform('mean')
                self.df_price[col] = self.df_price[col].fillna(group_mean)
            
            # 방법 5: 전체 평균
            overall_mean = self.df_price[col].mean()
            self.df_price[col] = self.df_price[col].fillna(overall_mean)
            
            filled = initial_missing - self.df_price[col].isna().sum()
            if filled > 0:
                print(f"  {col}: {filled}개 채움 (남은 결측치: {self.df_price[col].isna().sum()}개)")
    
    def fill_quantity_with_previous_day(self):
        """반입량 데이터: 전일 데이터로 채우기"""
        print("\n" + "=" * 60)
        print("반입량 데이터: 전일 데이터로 채우기")
        print("=" * 60)
        
        quantity_cols = [col for col in self.df_quantity.select_dtypes(include=[np.number]).columns 
                        if col not in ['is_weekend', 'is_holiday', 'is_trading_day']]
        group_cols = [col for col in self.df_quantity.columns 
                     if col not in quantity_cols and col not in ['DATE', 'is_weekend', 'is_holiday', 'is_trading_day']]
        
        # 0 값을 NaN으로 변환 (단, 합계나 총량 컬럼은 제외)
        for col in quantity_cols:
            zero_ratio = (self.df_quantity[col] == 0).sum() / len(self.df_quantity)
            # 0 값이 20% 이상이면 결측치로 간주 (단, 합계 컬럼은 제외)
            if zero_ratio > 0.2 and '총' not in col and '합계' not in col:
                self.df_quantity.loc[self.df_quantity[col] == 0, col] = np.nan
        
        # 그룹별로 전일 데이터로 채우기
        for col in quantity_cols:
            if col not in self.df_quantity.columns:
                continue
            
            initial_missing = self.df_quantity[col].isna().sum()
            if initial_missing == 0:
                continue
            
            if len(group_cols) > 0:
                self.df_quantity[col] = self.df_quantity.groupby(group_cols)[col].transform(
                    lambda x: x.ffill()
                )
            else:
                self.df_quantity[col] = self.df_quantity[col].ffill()
            
            filled = initial_missing - self.df_quantity[col].isna().sum()
            if filled > 0:
                print(f"  {col}: 전일 데이터로 {filled}개 채움")
    
    def fill_quantity_with_previous_year(self):
        """반입량 데이터: 전년 동일일자 데이터로 채우기"""
        print("\n" + "=" * 60)
        print("반입량 데이터: 전년 동일일자 데이터로 채우기")
        print("=" * 60)
        
        quantity_cols = [col for col in self.df_quantity.select_dtypes(include=[np.number]).columns 
                        if col not in ['is_weekend', 'is_holiday', 'is_trading_day']]
        group_cols = [col for col in self.df_quantity.columns 
                     if col not in quantity_cols and col not in ['DATE', 'is_weekend', 'is_holiday', 'is_trading_day']]
        
        for col in quantity_cols:
            initial_missing = self.df_quantity[col].isna().sum()
            if initial_missing == 0:
                continue
            
            filled_count = 0
            for idx, row in self.df_quantity[self.df_quantity[col].isna()].iterrows():
                current_date = row['DATE']
                prev_year_date = current_date - pd.DateOffset(years=1)
                
                if len(group_cols) > 0:
                    group_values = {g: row[g] for g in group_cols}
                    prev_year_data = self.df_quantity[
                        (self.df_quantity['DATE'] == prev_year_date) &
                        (self.df_quantity[col].notna())
                    ]
                    
                    for g in group_cols:
                        prev_year_data = prev_year_data[prev_year_data[g] == group_values[g]]
                    
                    if len(prev_year_data) > 0 and prev_year_data[col].iloc[0] > 0:
                        self.df_quantity.loc[idx, col] = prev_year_data[col].iloc[0]
                        filled_count += 1
                else:
                    prev_year_data = self.df_quantity[
                        (self.df_quantity['DATE'] == prev_year_date) &
                        (self.df_quantity[col].notna())
                    ]
                    if len(prev_year_data) > 0 and prev_year_data[col].iloc[0] > 0:
                        self.df_quantity.loc[idx, col] = prev_year_data[col].iloc[0]
                        filled_count += 1
            
            if filled_count > 0:
                print(f"  {col}: 전년 동일일자 데이터로 {filled_count}개 채움")
    
    def fill_quantity_holiday_values(self):
        """반입량 데이터: 휴일 처리 (선형보간)"""
        print("\n" + "=" * 60)
        print("반입량 데이터: 휴일 처리 (선형보간)")
        print("=" * 60)
        
        quantity_cols = [col for col in self.df_quantity.select_dtypes(include=[np.number]).columns 
                        if col not in ['is_weekend', 'is_holiday', 'is_trading_day']]
        group_cols = [col for col in self.df_quantity.columns 
                     if col not in quantity_cols and col not in ['DATE', 'is_weekend', 'is_holiday', 'is_trading_day']]
        
        for col in quantity_cols:
            holiday_mask = (~self.df_quantity['is_trading_day']) & (self.df_quantity[col].isna())
            initial_missing = holiday_mask.sum()
            
            if initial_missing == 0:
                continue
            
            # 그룹별로 선형보간 수행
            if len(group_cols) > 0:
                self.df_quantity[col] = self.df_quantity.groupby(group_cols)[col].transform(
                    lambda x: x.interpolate(method='linear', limit_direction='both')
                )
            else:
                self.df_quantity[col] = self.df_quantity[col].interpolate(method='linear', limit_direction='both')
            
            filled = initial_missing - self.df_quantity[col].isna().sum()
            if filled > 0:
                print(f"  {col}: 휴일 선형보간으로 {filled}개 채움")
    
    def fill_quantity_remaining_missing(self):
        """반입량 데이터: 남은 결측치 처리"""
        print("\n" + "=" * 60)
        print("반입량 데이터: 남은 결측치 처리")
        print("=" * 60)
        
        quantity_cols = [col for col in self.df_quantity.select_dtypes(include=[np.number]).columns 
                        if col not in ['is_weekend', 'is_holiday', 'is_trading_day']]
        group_cols = [col for col in self.df_quantity.columns 
                     if col not in quantity_cols and col not in ['DATE', 'is_weekend', 'is_holiday', 'is_trading_day']]
        
        for col in quantity_cols:
            if col not in self.df_quantity.columns:
                continue
                
            initial_missing = self.df_quantity[col].isna().sum()
            if initial_missing == 0:
                continue
            
            # 선형 보간
            if len(group_cols) > 0:
                self.df_quantity[col] = self.df_quantity.groupby(group_cols)[col].transform(
                    lambda x: x.interpolate(method='linear', limit_direction='both')
                )
            
            # 이동평균
            if len(group_cols) > 0:
                self.df_quantity[col] = self.df_quantity.groupby(group_cols)[col].transform(
                    lambda x: x.fillna(x.rolling(window=7, center=True, min_periods=1).mean())
                )
            
            # 같은 요일의 평균
            self.df_quantity['weekday'] = self.df_quantity['DATE'].dt.dayofweek
            if len(group_cols) > 0:
                weekday_mean = self.df_quantity.groupby(['weekday'] + group_cols)[col].transform('mean')
                self.df_quantity[col] = self.df_quantity[col].fillna(weekday_mean)
            else:
                weekday_mean = self.df_quantity.groupby('weekday')[col].transform('mean')
                self.df_quantity[col] = self.df_quantity[col].fillna(weekday_mean)
            self.df_quantity = self.df_quantity.drop('weekday', axis=1)
            
            # 그룹별 평균
            if len(group_cols) > 0:
                group_mean = self.df_quantity.groupby(group_cols)[col].transform('mean')
                self.df_quantity[col] = self.df_quantity[col].fillna(group_mean)
            
            # 전체 평균
            overall_mean = self.df_quantity[col].mean()
            self.df_quantity[col] = self.df_quantity[col].fillna(overall_mean)
            
            filled = initial_missing - self.df_quantity[col].isna().sum()
            if filled > 0:
                print(f"  {col}: {filled}개 채움 (남은 결측치: {self.df_quantity[col].isna().sum()}개)")
    
    def validate_data(self):
        """데이터 검증"""
        print("\n" + "=" * 60)
        print("데이터 검증")
        print("=" * 60)
        
        # 가격 데이터 검증
        price_cols = [col for col in self.df_price.select_dtypes(include=[np.number]).columns 
                     if col not in ['is_weekend', 'is_holiday', 'is_trading_day']]
        print("\n[가격 데이터]")
        for col in price_cols:
            missing = self.df_price[col].isna().sum()
            zeros = (self.df_price[col] == 0).sum()
            print(f"  {col}: 결측치 {missing}개, 0 값 {zeros}개")
            if len(self.df_price[col].dropna()) > 0:
                print(f"    범위: {self.df_price[col].min():.2f} ~ {self.df_price[col].max():.2f}")
        
        # 반입량 데이터 검증
        quantity_cols = [col for col in self.df_quantity.select_dtypes(include=[np.number]).columns 
                        if col not in ['is_weekend', 'is_holiday', 'is_trading_day']]
        print("\n[반입량 데이터]")
        for col in quantity_cols:
            missing = self.df_quantity[col].isna().sum()
            zeros = (self.df_quantity[col] == 0).sum()
            print(f"  {col}: 결측치 {missing}개, 0 값 {zeros}개")
            if len(self.df_quantity[col].dropna()) > 0:
                print(f"    범위: {self.df_quantity[col].min():.2f} ~ {self.df_quantity[col].max():.2f}")
    
    def save_processed_data(self, price_output='price_data_v1.xlsx', 
                           quantity_output='quantity_data_v1.xlsx'):
        """전처리된 데이터 저장 (플래그 컬럼 제거, 날짜 시간 부분 제거)"""
        print("\n" + "=" * 60)
        print("전처리된 데이터 저장")
        print("=" * 60)
        
        # 플래그 컬럼 제거
        df_price_save = self.df_price.drop(['is_weekend', 'is_holiday', 'is_trading_day'], axis=1, errors='ignore').copy()
        df_quantity_save = self.df_quantity.drop(['is_weekend', 'is_holiday', 'is_trading_day'], axis=1, errors='ignore').copy()
        
        # DATE 컬럼의 시간 부분 제거 (00:00:00 제거)
        if 'DATE' in df_price_save.columns:
            df_price_save['DATE'] = df_price_save['DATE'].dt.date
        if 'DATE' in df_quantity_save.columns:
            df_quantity_save['DATE'] = df_quantity_save['DATE'].dt.date
        
        # Excel 파일로 저장
        df_price_save.to_excel(price_output, index=False)
        df_quantity_save.to_excel(quantity_output, index=False)
        
        # 첫 행(헤더)을 텍스트 형식으로 설정
        if HAS_OPENPYXL:
            self._set_header_as_text(price_output)
            self._set_header_as_text(quantity_output)
        
        print(f"가격 데이터 저장: {price_output}")
        print(f"반입량 데이터 저장: {quantity_output}")
    
    def _set_header_as_text(self, file_path):
        """Excel 파일의 첫 행(헤더)을 텍스트 형식으로 설정"""
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 첫 행의 모든 셀을 텍스트 형식으로 설정
            for cell in ws[1]:
                cell.number_format = '@'  # 텍스트 형식
            
            wb.save(file_path)
            wb.close()
        except Exception as e:
            print(f"  경고: {file_path}의 헤더 형식 설정 중 오류 발생: {e}")
    
    def run_preprocessing(self):
        """전체 전처리 프로세스 실행"""
        print("\n" + "=" * 60)
        print("고급 데이터 전처리 시작")
        print("=" * 60)
        
        # 1. 데이터 로드
        self.load_data()
        
        # 2. 결측치 확인
        self.identify_missing_values()
        
        # 3. 가격 데이터 전처리
        self.fill_price_with_previous_day()      # 전일 데이터
        self.fill_price_with_previous_year()      # 전년 동일일자
        self.fill_price_holiday_values()         # 휴일 처리
        self.fill_price_remaining_missing()      # 남은 결측치
        
        # 4. 반입량 데이터 전처리
        self.fill_quantity_with_previous_day()  # 전일 데이터
        self.fill_quantity_with_previous_year()  # 전년 동일일자
        self.fill_quantity_holiday_values()      # 휴일 처리
        self.fill_quantity_remaining_missing()    # 남은 결측치
        
        # 5. 데이터 검증
        self.validate_data()
        
        # 6. 전처리된 데이터 저장
        self.save_processed_data()
        
        print("\n" + "=" * 60)
        print("데이터 전처리 완료!")
        print("=" * 60)
        
        return self.df_price, self.df_quantity


if __name__ == "__main__":
    # 전처리 실행
    preprocessor = AdvancedDataPreprocessor()
    df_price_processed, df_quantity_processed = preprocessor.run_preprocessing()
